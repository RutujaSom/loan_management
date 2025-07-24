import frappe
import csv
import io
from openpyxl import load_workbook
from datetime import datetime

@frappe.whitelist()
def import_loan_emi_from_file(file_url, file_type=None):
    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_content = file_doc.get_content()
    filename = file_doc.file_name.lower()

    # Validate file extension
    if file_type == "CSV" and not filename.endswith(".csv"):
        frappe.throw("Selected file type is CSV but uploaded file is not a .csv file.")
    elif file_type == "Excel" and not filename.endswith(".xlsx"):
        frappe.throw("Selected file type is Excel but uploaded file is not a .xlsx file.")

    imported_count = 0
    skipped_count = 0
    rows = []

    # Parse Excel
    if filename.endswith(".xlsx"):
        file_path = file_doc.get_full_path()
        with open(file_path, "rb") as f:
            workbook = load_workbook(f, data_only=True)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))

    # Parse CSV
    elif filename.endswith(".csv"):
        try:
            content = file_content.decode("utf-8")
        except UnicodeDecodeError:
            content = file_content.decode("latin-1")
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)

    for row in rows:
        member_no = (row.get("MEMBER NO") or "").strip()
        loan_id = (row.get("LOAN ID") or "").strip()
        emi_no = row.get("EMI NO")
        emi_date = parse_date(row.get("EMI DATE"))
        emi_day = (row.get("DAY") or "").strip()
        received_raw = row.get("RECEIVED DATE")
        received_date = parse_date(received_raw)
        emi_amount = row.get("EMI") or 0
        total = row.get("TOTAL") or 0
        remark = (row.get("REMARK IF ANY") or "").strip()

        cash_amount = row.get("CASH") or 0
        online_amount_1 = row.get("ONLINE") or 0
        utr_no_1 = (row.get("UTR NO") or "")
        online_amount_2 = row.get("ONLINE 1") or 0
        utr_no_2 = (row.get("UTR NO 1") or "")
        print('remark ......',remark , ' loan_id .... ',loan_id,'emi_no ....',emi_no)

        if not member_no or not loan_id:
            skipped_count += 1
            continue

        # # Validate Customer
        customer = frappe.db.get_value("Customer", {"member_no": member_no}, "name")
        if not customer:
            skipped_count += 1
            continue

        # # Validate Loan
        loan = frappe.db.get_value("Customer Loan", {"loan_id": loan_id, "customer": customer}, "name")
        if not loan:
            skipped_count += 1
            continue

        # # Check for duplicate loan + emi_no
        if frappe.db.exists("Loan EMI", {"loan": loan, "emi_no": emi_no}):
            skipped_count += 1
            continue


        # Create Loan EMI doc
        loan_emi = frappe.new_doc("Loan EMI")
        loan_emi.customer = customer
        loan_emi.loan = loan
        loan_emi.emi_amount = emi_amount
        loan_emi.emi_no = emi_no
        loan_emi.emi_date = emi_date
        loan_emi.emi_day = emi_day
        loan_emi.total = total

        # Handle emi_complete flag and received_date
        if isinstance(received_raw, str) and received_raw.strip().lower() == "nil":
            loan_emi.received_date = None
            loan_emi.emi_complete = False
        elif received_date:
            loan_emi.received_date = received_date
            loan_emi.emi_complete = True
        else:
            loan_emi.emi_complete = False

        # Add EMI Transaction rows
        if cash_amount:
            loan_emi.append("emi_transaction", {
                "payment_mode": "Cash",
                "amount": cash_amount,
                "utr_no": "",
                "remark": remark
            })

        if online_amount_1:
            loan_emi.append("emi_transaction", {
                "payment_mode": "Online",
                "amount": online_amount_1,
                "utr_no": utr_no_1,
                "remark": remark
            })

        if online_amount_2:
            loan_emi.append("emi_transaction", {
                "payment_mode": "Online",
                "amount": online_amount_2,
                "utr_no": utr_no_2,
                "remark": remark
            })

        loan_emi.insert(ignore_permissions=True)
        imported_count += 1

    return f"{imported_count} EMI record(s) imported. {skipped_count} skipped due to missing or invalid customer/loan data."


# Helper: Parse flexible date formats
def parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, str):
        value = value.strip()
        formats = [
            "%d-%m-%Y", "%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y",
            "%d %B %Y", "%b %d, %Y", "%B %d, %Y", "%d.%m.%Y"
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).date().isoformat()
            except ValueError:
                continue
    return None






import frappe
from datetime import datetime, timedelta


@frappe.whitelist()
def fetch_filtered_emis(customer=None, days_ahead=1):
    days_ahead = int(days_ahead)
    end_date = (datetime.now() + timedelta(days=days_ahead)).strftime("%Y-%m-%d")

    conditions = "1=1"
    # if customer:
    #     conditions += f" AND emi.customer = {frappe.db.escape(customer)}"
    conditions += f" AND emi.emi_date = {frappe.db.escape(end_date)}"
    
    data = frappe.db.sql("""
        SELECT
            emi.name,
            emi.loan,
            emi.customer,
            emi.emi_no,
            emi.emi_date,
            emi.emi_amount,
            loan.loan_id,
            cust.title,
            cust.mobile_no
        FROM `tabLoan EMI` emi
        LEFT JOIN `tabCustomer Loan` loan ON loan.name = emi.loan
        LEFT JOIN `tabCustomer` cust ON cust.name = emi.customer
        WHERE {conditions}
        ORDER BY emi.emi_no
    """.format(conditions=conditions), as_dict=True)
    print('data ....', data)
    return data


