import frappe
import csv
import io
from openpyxl import load_workbook

@frappe.whitelist()
def import_customer_loan_from_file(file_url, test_plan=None, file_type=None):
    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_content = file_doc.get_content()
    filename = file_doc.file_name.lower()

    # Validate file extension based on selected file_type
    if file_type == "CSV" and not filename.endswith(".csv"):
        frappe.throw("Selected file type is CSV but uploaded file is not a .csv file.")
    elif file_type == "Excel" and not filename.endswith(".xlsx"):
        frappe.throw("Selected file type is Excel but uploaded file is not a .xlsx file.")

    imported_count = 0
    skipped_count = 0
    rows = []

    # Parse Excel
    if filename.endswith(".xlsx"):
        file_path = file_doc.get_full_path()
        with open(file_path, "rb") as f:
            workbook = load_workbook(f, data_only=True)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))

    # Parse CSV
    elif filename.endswith(".csv"):
        try:
            content = file_content.decode("utf-8")
        except UnicodeDecodeError:
            content = file_content.decode("latin-1")
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)

    # Process each row
    for row in rows:
        loan_id = (row.get("LOAN ID") or "").strip()
        amount = (str(row.get("LOAN") or ""))
        member_no = (row.get("MEMBER NO") or "").strip()
        terms = (row.get("TERMS") or "")
        emi_amount = (row.get("EMI") or "")
        date_of_sanction = (row.get("DATE OF SANCTION") or "")
        date_of_sanction = parse_date(date_of_sanction)
        print('date_of_sanction ....',date_of_sanction,"....", type(date_of_sanction),'member_no ....',member_no)

        # Skip if member_no or loan_id is missing
        if not member_no or not loan_id:
            skipped_count += 1
            continue

        # Check if customer exists
        customer_doc_name = frappe.db.get_value("Customer", {"member_no": member_no}, "name")
        if not customer_doc_name:
            skipped_count += 1
            continue

        # Check if loan already exists
        if frappe.db.exists("Customer Loan", {"loan_id": loan_id}):
            skipped_count += 1
            continue

        # Create new Customer Loan record
        customer_loan = frappe.new_doc("Customer Loan")
        customer_loan.loan_id = loan_id
        customer_loan.date_of_sanction = date_of_sanction
        customer_loan.terms = terms
        customer_loan.amount = amount
        customer_loan.emi_amount = emi_amount
        customer_loan.customer = customer_doc_name  # Link to Customer

        # Save Loan
        customer_loan.insert(ignore_permissions=True)
        imported_count += 1

    return f"{imported_count} Loan(s) imported. {skipped_count} record(s) skipped due to missing data, duplicate member numbers, or duplicate loan IDs."



from datetime import datetime

def parse_date(value):
    if not value:
        return None

    # If already a datetime object (like from Excel), return ISO
    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, str):
        value = value.strip()

        # List of possible formats to try
        date_formats = [
            "%d-%m-%Y", "%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y",
            "%d %B %Y", "%b %d, %Y", "%B %d, %Y", "%d.%m.%Y"
        ]

        for fmt in date_formats:
            try:
                return datetime.strptime(value, fmt).date().isoformat()
            except ValueError:
                continue

    return None  # Unrecognized or invalid format
