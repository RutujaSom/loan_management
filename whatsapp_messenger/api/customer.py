import frappe
import csv
import io
from openpyxl import load_workbook


@frappe.whitelist()
def import_customers_from_file(file_url, test_plan=None, file_type=None):
    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_content = file_doc.get_content()
    filename = file_doc.file_name.lower()

    # Validate file extension based on selected file_type
    if file_type == "CSV" and not filename.endswith(".csv"):
        frappe.throw("Selected file type is CSV but uploaded file is not a .csv file.")
    elif file_type == "Excel" and not filename.endswith(".xlsx"):
        frappe.throw("Selected file type is Excel but uploaded file is not a .xlsx file.")

    imported_count = 0
    skipped_count = 0
    rows = []
    total_customers = frappe.db.count("Customer")

    frappe.throw(total_customers)

    # Parse Excel
    if filename.endswith(".xlsx"):
        file_path = file_doc.get_full_path()
        with open(file_path, "rb") as f:
            workbook = load_workbook(f, data_only=True)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))

    # Parse CSV
    elif filename.endswith(".csv"):
        try:
            content = file_content.decode("utf-8")
        except UnicodeDecodeError:
            content = file_content.decode("latin-1")
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)

    # Process each row
    for row in rows:
        group_id = (row.get("GROUP CODE") or "").strip()
        title = (row.get("NAME OF MEMBER") or "").strip()
        mobile_no = (str(row.get("MOB") or "")).strip()
        member_no = (row.get("MEMBER NO") or "").strip()

        # Skip if member_no is missing
        if not member_no:
            skipped_count += 1
            continue

        # Check for duplicate member_no
        if frappe.db.exists("Customer", {"member_no": member_no}):
            skipped_count += 1
            continue

        # Create new Customer record
        customer = frappe.new_doc("Customer")
        customer.group_id = group_id
        customer.member_no = member_no
        customer.title = title
        customer.mobile_no = mobile_no

        # Save Customer
        customer.insert(ignore_permissions=True)
        imported_count += 1

    return f"{imported_count} Customer(s) imported. {skipped_count} record(s) skipped due to duplicate member numbers."
